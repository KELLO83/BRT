import cv2
import numpy as np
from glob import glob
from natsort import natsorted
import os
import matplotlib.pyplot as plt
from tqdm import tqdm
from scipy.stats import norm
import logging
import time

logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")

stream_handler = logging.StreamHandler()
stream_handler.setFormatter(formatter)
logger.addHandler(stream_handler)

image_number = 24
os.environ["QT_DEBUG_PLUGINS"] = "0"
seat_file_list = natsorted(
    glob(
        os.path.join(f"cv2_diff_test/raw_seat/{image_number}", "*.jpg"), recursive=True
    )
)
unseat_file_list = natsorted(
    glob(
        os.path.join(f"cv2_diff_test/removed_raw_seat/{image_number}", "*.jpg"),
        recursive=True,
    )
)


if image_number <= 10:
    empty_image = cv2.imread("cv2_diff_test/front_empty.jpg", cv2.IMREAD_COLOR)
elif image_number <= 15:
    empty_image = cv2.imread("cv2_diff_test/front2_empty.jpg", cv2.IMREAD_COLOR)
elif image_number <= 24:
    empty_image = cv2.imread("cv2_diff_test/back_empty.jpg", cv2.IMREAD_COLOR)
else:
    raise IndexError

sheet_no = f"{image_number}"

sheet_cordinate = {
    4: [1193, 492, 1346, 717],
    8: [1104, 333, 1181, 437],
    11: [697, 220, 801, 394],
    12 : [589, 253, 686, 385],
    24 : [1098 , 274 ,1166, 364 ],
}


def image_analyze(file_list: list, empty: np.ndarray):
    full_h, full_w, _ = cv2.imread(file_list[0]).shape

    x1, y1, x2, y2 = sheet_cordinate[image_number]

    # cv2.namedWindow("t", cv2.WINDOW_NORMAL)
    # cv2.imshow("t", empty)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()

    empty = empty[y1:y2, x1:x2]

    # cv2.namedWindow("t", cv2.WINDOW_NORMAL)
    # cv2.imshow("t", empty)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    h, w, _ = empty.shape

    red_score = np.zeros((h, w), dtype=np.float128)
    blue_score = np.zeros((h, w), dtype=np.float128)
    green_score = np.zeros((h, w), dtype=np.float128)

    for i in tqdm(file_list):
        image = cv2.imread(i)
        image = image[y1:y2, x1:x2]

        diff = cv2.absdiff(empty, image)

        # cv2.namedWindow("c",cv2.WINDOW_NORMAL)
        # cv2.imshow("c" , diff)
        # cv2.waitKey(0)

        diff = diff.astype(np.float64)
        blue_score += diff[:, :, 0]
        green_score += diff[:, :, 1]
        red_score += diff[:, :, 2]

    red_score = (red_score / len(file_list)).astype(np.uint8)
    blue_score = (blue_score / len(file_list)).astype(np.uint8)
    green_score = (green_score / len(file_list)).astype(np.uint8)

    merge = cv2.merge([blue_score, green_score, red_score])

    merge = np.clip(merge, 0, 255)

    # cv2.namedWindow("m",cv2.WINDOW_NORMAL)
    # cv2.imshow("m" , merge)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()

    b, g, r = cv2.split(merge)

    def hist__(b, g, r):
        b_hist = cv2.calcHist([b], [0], None, [256], [0, 256])
        g_hist = cv2.calcHist([g], [0], None, [256], [0, 256])
        r_hist = cv2.calcHist([r], [0], None, [256], [0, 256])

        def calc_hist_percentiles(
            hist, color, subplot_position, percentiles=[0.75, 0.8, 0.9]
        ):
            cdf = np.cumsum(hist)
            total_pixels = cdf[-1]
            p_index = []
            for p in percentiles:
                p_value = total_pixels * p
                idx = np.where(cdf >= p_value)[0][0]
                # print(idx)
                p_index.append(idx)

            plt.subplot(subplot_position)
            plt.plot(hist, color=color)

            color = ["b", "g", "r", "c", "m", "y"]
            for i, idx in enumerate(p_index):
                plt.axvline(
                    idx,
                    linestyle="--",
                    label=f"{int(percentiles[i] * 100)}% at {idx}",
                    color=color[i % len(color)],
                )
                plt.text(
                    idx,
                    max(hist) * (0.9 - i * 0.2),
                    f"{int(percentiles[i] * 100)}%: {idx}",
                    color=color[i % len(color)],
                    ha="center",
                )

            plt.legend()

            return p_index

        plt.figure(figsize=(12, 12))
        b_idx = calc_hist_percentiles(b_hist, "b", 221)
        g_idx = calc_hist_percentiles(g_hist, "g", 222)
        r_idx = calc_hist_percentiles(r_hist, "r", 223)
        # plt.show()

        return b_idx, g_idx, r_idx

    RGB_th_buffer = []

    RGB_th_buffer = zip(*hist__(b, g, r))

    for index, i in enumerate(RGB_th_buffer):
        b_t, r_t, g_t = i

        background_and = np.zeros((full_h, full_w), dtype=np.uint8)
        background_or = np.zeros((full_h, full_w), dtype=np.uint8)

        b_mask = b > b_t
        g_mask = g > g_t
        r_mask = r > r_t

        combined_and_mask = b_mask & g_mask & r_mask
        combined_or_mask = b_mask | g_mask | r_mask

        background_and[y1:y2, x1:x2] = combined_and_mask
        background_or[y1:y2, x1:x2] = combined_or_mask

        Kernel_density_estimation(index, background_and, i, True)
        Kernel_density_estimation(index, background_or, i, False)
    
    import pandas as pd
    file_path = "cv2_diff_test/rs.xlsx"
    df = pd.read_excel(f'{file_path}')
    df['Rank'] = df['Accuracy'].rank(ascending=False, method='min').astype(int)
    print(df)
    df.to_excel('ranked_file.xlsx', index=False)

def Kernel_density_estimation(index, mask, color_th, flag):
    b_t, r_t, g_t = color_th

    if flag == True:
        OPERATION = "AND"
    else:
        OPERATION = "OR"

    global empty_image, seat_file_list, unseat_file_list
    import pandas as pd

    percentlines = [0.75, 0.8, 0.9]

    acc = []
    mask_cordinate = np.where(mask >= 1)
    mask_cordinate = list(zip(mask_cordinate[0], mask_cordinate[1]))
    print(
        "============================ Run Kernel_Density_Estimation ========================="
    )
    for image_path in tqdm(seat_file_list):
        pixel_count = 0

        compare = cv2.imread(image_path)
        diff = cv2.absdiff(empty_image, compare)

        b_d, g_d, r_d = cv2.split(diff)

        for i in mask_cordinate:
            y, x = i

            b = b_d[y][x]
            g = g_d[y][x]
            r = r_d[y][x]

            if flag == True:
                if b >= b_t and g >= g_t and r > r_t:
                    pixel_count += 1

            else:
                if b >= b_t or g >= g_t or r > r_t:
                    pixel_count += 1

        acc.append(pixel_count)

    acc = np.array(acc).astype(np.uint32)
    acc = acc[~np.isnan(acc)]  # NaN 제거
    acc = acc[np.isfinite(acc)]  # Infinite 값 제거
    from scipy.stats import gaussian_kde

    density = gaussian_kde(acc)
    xs = np.linspace(min(acc), max(acc), 1000)
    density_value = density(xs)
    for i in range(len(xs) - 1):
        if density_value[i] <= 0.001:  # 밀도가 0에 가까운 구간
            plt.plot(xs[i : i + 2], density_value[i : i + 2], color="blue")
        else:
            plt.plot(xs[i : i + 2], density_value[i : i + 2], color="red")

    CDF = np.cumsum(density_value)
    CDF = CDF / CDF[-1]

    from scipy.interpolate import interp1d

    INV_CDF = interp1d(CDF, xs)

    p70 = INV_CDF(0.70)  # 70% 구간
    p80 = INV_CDF(0.80)  # 80% 구간
    p90 = INV_CDF(0.90)  # 90% 구간
    p99 = INV_CDF(0.99)  # 99% 구간
    p100 = INV_CDF(1.00)  # 100% 구간

    plt.axvline(p100, color="b", linestyle="--", label=f"100% at x={p100:.2f}")

    plt.text(p100, max(density_value), f"100%: {p100:.2f}", color="b", ha="center")

    plt.axvline(p70, color="g", linestyle="--", label=f"70% at x={p70:.2f}")
    plt.axvline(p80, color="y", linestyle="--", label=f"80% at x={p80:.2f}")
    plt.axvline(p90, color="c", linestyle="--", label=f"90% at x={p90:.2f}")
    plt.axvline(p99, color="m", linestyle="--", label=f"99% at x={p99:.2f}")

    plt.text(p70, max(density_value) * 0.6, f"70%: {p70:.2f}", color="g", ha="center")
    plt.text(p80, max(density_value) * 0.7, f"80%: {p80:.2f}", color="y", ha="center")
    plt.text(p90, max(density_value) * 0.8, f"90%: {p90:.2f}", color="c", ha="center")
    plt.text(p99, max(density_value) * 0.9, f"99%: {p99:.2f}", color="m", ha="center")

    p0 = xs[0]
    p1 = INV_CDF(0.01)  # 1% 구간
    p3 = INV_CDF(0.03)  # 3% 구간
    p5 = INV_CDF(0.05)  # 5% 구간
    p10 = INV_CDF(0.10)  # 10% 구간

    plt.axvline(p0, color="r", linestyle="--", label=f"0% at x={p0:.2f}")
    plt.axvline(p1, color="g", linestyle="--", label=f"1% at x={p1:.2f}")
    plt.axvline(p3, color="y", linestyle="--", label=f"3% at x={p3:.2f}")
    plt.axvline(p5, color="c", linestyle="--", label=f"5% at x={p5:.2f}")
    plt.axvline(p10, color="m", linestyle="--", label=f"10% at x={p10:.2f}")

    # 수직선에 텍스트 추가
    plt.text(p0, max(density_value) * 0.5, f"0%: {p0:.2f}", color="r", ha="center")
    plt.text(p1, max(density_value) * 0.6, f"1%: {p1:.2f}", color="g", ha="center")
    plt.text(p3, max(density_value) * 0.7, f"3%: {p3:.2f}", color="y", ha="center")
    plt.text(p5, max(density_value) * 0.8, f"5%: {p5:.2f}", color="c", ha="center")
    plt.text(p10, max(density_value) * 0.9, f"10%: {p10:.2f}", color="m", ha="center")

    plt.title(f"{percentlines[index]}% {OPERATION} Excute")
    plt.xlabel("값")
    plt.ylabel("빈도")
    # plt.show()

    PIXEL_TH = {"1%": p1, "3%": p3, "5%": p5}

    for STR_PT, PT in PIXEL_TH.items():
        stop_point = PT  # 착석하지않은사람이 p1보다는 커야하고
        print("{}% 임계점  : {:.2f} / {:.2f}".format(STR_PT, stop_point, p100))

        undetection_image = []

        for i in tqdm(seat_file_list):
            image_name = i
            pixel_count = 0

            compare = cv2.imread(i)
            compare_copy = compare.copy()
            diff = cv2.absdiff(empty_image, compare)

            b_d, g_d, r_d = cv2.split(diff)
            for i in mask_cordinate:
                y, x = i
                compare_copy[y][x] = (0, 0, 255)
                b = b_d[y][x]
                g = g_d[y][x]
                r = r_d[y][x]

                if flag == True:
                    if b >= b_t and g >= g_t and r > r_t:
                        pixel_count += 1

                else:
                    if b >= b_t or g >= g_t or r > r_t:
                        pixel_count += 1

            # print(f" {pixel_count} / {len(mask_cordinate)}")
            # cv2.namedWindow('c',cv2.WINDOW_NORMAL)
            # cv2.imshow("c",compare_copy)
            # cv2.waitKey(0)

            if pixel_count <= stop_point:  # 미탐지율
                undetection_image.append(os.path.basename(image_name))
                # cv2.namedWindow('c',cv2.WINDOW_NORMAL)
                # cv2.imshow("c",compare_copy)
                # cv2.waitKey(0)

        false_alarm = []
        for i in tqdm(unseat_file_list):
            image_name = i
            pixel_count = 0

            compare = cv2.imread(i)
            diff = cv2.absdiff(empty_image, compare)

            b_d, g_d, r_d = cv2.split(diff)

            for i in mask_cordinate:
                y, x = i

                b = b_d[y][x]
                g = g_d[y][x]
                r = r_d[y][x]

                if flag == True:
                    if b >= b_t and g >= g_t and r > r_t:
                        pixel_count += 1

                else:
                    if b >= b_t or g >= g_t or r > r_t:
                        pixel_count += 1

            if pixel_count >= stop_point:  # 오경보
                false_alarm.append(os.path.basename(image_name))

        TP , TN , FP , FN , Rc , Ac,  F1, ACC = Performance_Metrix(undetection_image, false_alarm)

        T = time.strftime("%H:%M:%S")

        logging.info(T)

        print(
            "======================================================================================"
        )
        print(
            f"{percentlines[index]} OP : {OPERATION} ---> Undetection Image {len(undetection_image)} / {len(seat_file_list)}\
            Failed Alarm {len(false_alarm)} / {len(unseat_file_list)}"
        )
        print(
            f"미탐지율 : {len(undetection_image) / len(seat_file_list) * 100 : .2f}%\
            오경보율 : {len(false_alarm) / len(unseat_file_list) * 100 : .2f}%  F1-SCORE : {F1 : .4f} ACC  : {ACC : .4f}"
        )
        print(
            "======================================================================================"
        )

        stop_point = round(float(stop_point) , 3)
        Data = [
            OPERATION ,
            STR_PT ,
            stop_point,
            percentlines[index],
            color_th[0],
            color_th[1],
            color_th[2],
            TP,
            TN,
            FP,
            FN,
            Rc,
            Ac,
            F1,
            ACC
        ]
        # data = [
        #     OPERATION,
        #     STR_PT,
        #     percentlines[index],
        #     color_th[0],
        #     color_th[1],
        #     color_th[2],
        #     f"{len(undetection_image)} / {len(seat_file_list)}",
        #     f"{len(false_alarm)} / {len(unseat_file_list)} ",
        #     f"{100 - (len(undetection_image) / len(seat_file_list)) * 100  : .1f}%",
        #     f"{len(false_alarm) / len(unseat_file_list) * 100 : .1f}%",
        #     round(F1, 5),
        #     round(ACC, 5),
        # ]

        append_data_to_excel(Data)


def Performance_Metrix(undetection_image, False_Alarm):
    global seat_file_list, unseat_file_list

    s_len = len(seat_file_list)
    uns_len = len(unseat_file_list)

    TP = s_len - len(undetection_image)
    TN = uns_len - len(False_Alarm)
    FP = len(False_Alarm)
    FN = len(undetection_image)

    Recall = TP / (TP + FN)
    Precision = TP / (TP + FP)

    F1_SCORE = 2 * (Precision * Recall) / (Precision + Recall)
    Accuracy = (TP + TN) / (TP + TN + FP + FN)

    result = list(
        map(
            lambda x: round(x, 3) if not isinstance(x, int) else x,
            [TP, TN, FP, FN, Recall, Precision, F1_SCORE, Accuracy],
        )
    )

    return result

def append_data_to_excel(data, sheet_name=f"Sheet{sheet_no}"):
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd

    print(data)
    df = pd.DataFrame(
        [data],
        columns=[
            "Operation",
            "Pixel Threshold",
            "Pixel TH value",
            "Mask Threshold",
            "Blue",
            "Green",
            "Red",
            "TP",
            "TN",
            "FP",
            "FN",
            "Recall",
            "Precsion",
            "F1-SCORE",
            "Accuracy",
        ],
    )
    file_path = "cv2_diff_test/rs.xlsx"
    try:
        excel_file = openpyxl.load_workbook(file_path)
        if sheet_name in excel_file.sheetnames:
            excel_ws = excel_file[sheet_name]
        else:
            excel_ws = excel_file.create_sheet(sheet_name)
            excel_ws.append(df.columns.tolist())

    except FileNotFoundError:
        wb = Workbook()
        excel_ws = wb.active
        excel_ws.title = sheet_name

        excel_ws.append(df.columns.tolist())

        wb.save(file_path)
        excel_file = openpyxl.load_workbook(file_path)
        excel_ws = excel_file[sheet_name]

    for r in df.itertuples(index=False):
        excel_ws.append(r)

    excel_file.save(file_path)


if __name__ == "__main__":
    image_analyze(seat_file_list, empty_image)
