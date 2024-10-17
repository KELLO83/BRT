import os
import cv2
import numpy as np
from ultralytics import YOLO
from tqdm.auto import  tqdm
import os
import fish_map
import box_module
import logging
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%H:%M:%S'
)

logging.info("Yolo Detect run ...")

import time
import functools
def timing_decorator(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        current_alpha = args[1]
        print("Model : ",args[0].model.model_name)
        print("ALPHA: ", current_alpha)
        print(f"Function '{func.__name__}' started at {time.strftime('%H:%M:%S', time.localtime(start_time))}")
        
        result = func(*args, **kwargs)
        
        end_time = time.time()
        print(f"Function '{func.__name__}' ended at {time.strftime('%H:%M:%S', time.localtime(end_time))}")
        print(f"Total execution time: {end_time - start_time:.4f} seconds")

        return result
    return wrapper


class YOLODetector:
    def __init__(self, f1 , number , alpha , box_p ,model_path='yolo8x.pt'):
        self.number = number
        self.distortion_list = f1
        self.model = YOLO(model_path).to("cuda")
        self.empty_count = 0
        self.false_alarm = 0
        self.not_same_source = 0

        self.alpha = alpha

        self.box_people = box_p


        self.PerformanceEvaluator = PerformanceEvaluator()


    @property
    def run_all(self):
        for alpha_value in self.alpha:
            self.run(alpha_value)
        

    def nmx_box_to_cv2_loc(self , boxes):
        x1 , y1 , w, h = boxes
        x2 = x1 + w
        y2 = y1 + h

        return [x1 , y1 , x2 , y2]

    def apply_nms(self, boxes, scores, iou_threshold=0.6):
        indices = cv2.dnn.NMSBoxes(boxes, scores, 0.5, iou_threshold)
        if isinstance(indices, list) and len(indices) > 0:
            return [boxes[i[0]] for i in indices]
        elif isinstance(indices, np.ndarray) and indices.size > 0:
            return [boxes[i] for i in indices.flatten()]
        else:
            return []
        
    @timing_decorator
    def run(self , alpha_value ):
        for iterable_count , i in tqdm(enumerate(self.distortion_list),total=len(self.distortion_list)):
            image_name = i
            distort_image   = cv2.imread(i , cv2.IMREAD_COLOR)

            if distort_image is not None:
                h, w, _ = list(map(int, distort_image.shape))

                # Add black padding around the distorted image
                black = np.zeros(((int(w - h) // 2), w, 3), dtype=np.uint8)
                frame_new = cv2.vconcat([black, distort_image])
                frame_new = cv2.vconcat([frame_new, black])


                # Recalculate the height and width after padding
                h, w, _ = list(map(int, frame_new.shape))

                # Apply fisheye_to_plane_info mapping on the newly padded image
                undisort_image = np.array(fish_map.fisheye_to_plane_info(frame_new, h, w, 180, 90, 640, 0, 0))
                undisort_image = cv2.resize(undisort_image, (640, 640), interpolation=cv2.INTER_LANCZOS4)
            

            result = self.model(undisort_image , verbose = False , classes = 0)

            boxes = []
            scores = []

            for res in result:
                for box in res.boxes:
                    if int(box.cls) == 0:
                        x1 , y1 , w , h = box.xyxy[0].tolist()
                        score = box.conf
                        boxes.append([int(x1), int(y1), int(w - x1), int(h - y1)])
                        scores.append(float(score.detach().cpu().numpy()))

            nmx_boxes = self.apply_nms(boxes , scores)
            nmx_boxes = list(map(self.nmx_box_to_cv2_loc , nmx_boxes))
            if not nmx_boxes:
                self.empty_count +=1
                continue


            answer , answer_location , ans_img =  box_module.second(nmx_boxes , undisort_image.copy() , Answer= True , ALPHA = 1 )
            compare_match , compare_location  , compare_img = box_module.second(nmx_boxes , undisort_image.copy() , Answer= False, ALPHA = alpha_value)

            self.PerformanceEvaluator.Metrix(answer_location , compare_location  , alpha_value)

 
            answer_location = dict(sorted(answer_location.items()))
            compare_location = dict(sorted(compare_location.items()))

            if answer != list(answer_location.keys()) or compare_match != list(compare_location.keys()):
                print(answer)
                print(answer_location)
                print(compare_match)
                print(compare_location)
                raise ValueError
            

            stop_point = False
            if len(answer_location) == len(compare_location):  # 길이가 같은 경우
                for answer_key in answer_location:
                    if answer_key not in compare_location:  # 키가 다르면 오경보 처리
                        self.false_alarm += 1
                        stop_point = True
                        break
                    else:  # 키가 같다면 값 비교
                        if answer_location[answer_key] != compare_location[answer_key]:
                            self.not_same_source += 1
                            stop_point = True
                            break
            else:  # 길이가 다르면 오경보 처리
                self.false_alarm += 1
                stop_point = True

            if stop_point:
                print('===================================================')
                print(f"{image_name}")
                print(f"{answer_location}")
                print(f"{compare_location}")
                print(f"오경보 : {self.false_alarm}")
                print(f"출처다름 : {self.not_same_source}")
            

        # cv2.namedWindow("Answer" , cv2.WINDOW_NORMAL)
        # cv2.namedWindow("compare", cv2.WINDOW_NORMAL)
        # cv2.imshow("Answer" , ans_img)
        # cv2.imshow("compare",compare_img)
        # cv2.waitKey(0)
        #cv2.destroyAllWindows()

        self.PerformanceEvaluator.Performeance_Metrix(alpha = alpha_value)
        self.PerformanceEvaluator.write_Q_to_excel()


import pandas as pd

class PerformanceEvaluator:
    def __init__(self):
        # 각 인원 수에 대한 TP, TN, FN, FP를 누적하기 위한 딕셔너리
        self.record = {}
        # 좌석 비율 설정: 인원 수를 키로 하고, 해당 좌석 번호의 집합을 값으로 가집니다.
        self.sit_proportion = {
            2 : {10,13},
            4 : {7, 8, 11, 12},
            6: {7, 8, 10, 11, 12, 13},
            9: {5,6,7,8,9,10,11,12,13},
            13: {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13}
        }
        # 성능 메트릭을 저장하기 위한 딕셔너리
        self.performance_dict = {}

        self.Q = {}

        self.previous_alpha = None

    def append_data_to_excel(self, dataFrame, sheet_name):
        import openpyxl
        from openpyxl import load_workbook
        import pandas as pd
        import os

        file_path = 'fit_sit.xlsx'

        # 파일이 존재하는지 확인
        if os.path.exists(file_path):
            # 엑셀 파일 로드
            excel_file = openpyxl.load_workbook(file_path)
        else:
            # 새로운 워크북 생성
            excel_file = openpyxl.Workbook()
            # 기본 시트 제거
            default_sheet = excel_file.active
            excel_file.remove(default_sheet)

        # 시트가 이미 존재하는지 확인
        if sheet_name in excel_file.sheetnames:
            excel_ws = excel_file[sheet_name]
            # 시트가 비어 있으면 헤더 추가
            if excel_ws.max_row == 1 and excel_ws.max_column == 1 and excel_ws.cell(1,1).value is None:
                excel_ws.append(dataFrame.columns.tolist())
        else:
            # 새로운 시트 생성
            excel_ws = excel_file.create_sheet(sheet_name)
            # 헤더 추가
            excel_ws.append(dataFrame.columns.tolist())

        # 데이터 추가
        for r in dataFrame.itertuples(index=False, name=None):
            excel_ws.append(r)

        # 엑셀 파일 저장
        excel_file.save(file_path)

    def write_Q_to_excel(self):
        for num_people, result in self.Q.items():
            # 결과 딕셔너리를 데이터프레임으로 변환
            df = pd.DataFrame([result])
            # 시트 이름 생성
            sheet_name = f"sheet_new_{num_people}"
            # 데이터프레임을 엑셀에 추가
            self.append_data_to_excel(df, sheet_name)



    def Metrix(self, answer_location, compare_location , alpha ):

        if self.previous_alpha != alpha:
            # 알파 값이 변경되었으므로 self.record 초기화
            self.record = {}
            # self.Q도 초기화해야 할 경우 여기에 추가
            self.Q = {}
            # previous_alpha 업데이트
            self.previous_alpha = alpha

        # mapper 생성: 각 좌석에 대한 TP, TN, FN, FP 값을 저장
        mapper = [None for _ in range(13)]
        ans_key = [k - 1 for k in answer_location.keys()]
        compare_key = [k - 1 for k in compare_location.keys()]

        # 기본적인 TP, TN, FN, FP 할당
        for i in range(13):
            if i in ans_key and i in compare_key:
                mapper[i] = 'TP'
            elif i not in ans_key and i not in compare_key:
                mapper[i] = 'TN'
            elif i in ans_key and i not in compare_key:
                mapper[i] = 'FN'
            elif i not in ans_key and i in compare_key:
                mapper[i] = 'FP'

        # 비교 위치의 값이 정답 위치의 값과 다른 경우 FP로 처리
        for key, value in compare_location.items():
            adjust_key = key - 1
            try:
                answer_value = answer_location[key]
                if value != answer_value:
                    mapper[adjust_key] = 'FP'
            except KeyError:
                if adjust_key not in ans_key:
                    mapper[adjust_key] = 'FP'

        # 각 인원 수에 대해 메트릭 계산 및 누적
        for num_people, seats_set in self.sit_proportion.items():
            # 좌석 번호를 인덱스로 변환
            mapping_ = [seat - 1 for seat in seats_set]
            TP = TN = FN = FP = 0

            # 현재 좌석 그룹에 대한 메트릭 계산
            for idx in mapping_:
                val = mapper[idx]
                if val == 'TP':
                    TP += 1
                elif val == 'TN':
                    TN += 1
                elif val == 'FN':
                    FN += 1
                elif val == 'FP':
                    FP += 1

            # 기존 값 가져오기 (딕셔너리 형태)
            prev_counts = self.record.get(num_people, {'TP': 0, 'TN': 0, 'FN': 0, 'FP': 0})
            # 새로운 값과 합산하여 누적
            self.record[num_people] = {
                'TP': prev_counts['TP'] + TP,
                'TN': prev_counts['TN'] + TN,
                'FN': prev_counts['FN'] + FN,
                'FP': prev_counts['FP'] + FP
            }

            # 총합 검증 (선택 사항)
            total = TP + TN + FN + FP
            if total != len(mapping_):
                print(f"인원 수 {num_people}명에 대한 좌석 수 불일치: TP={TP}, TN={TN}, FN={FN}, FP={FP}")
                raise ValueError(f"총합이 좌석 수 {len(mapping_)}와 일치하지 않습니다.")

    def Performeance_Metrix(self, alpha):
        # 결과를 저장할 리스트 초기화
        data_list = []

        # self.Q가 초기화되어 있는지 확인
        if not hasattr(self, 'Q'):
            self.Q = {}

        # self.record의 각 키(인원 수)에 대해 반복
        for num_people, counts in self.record.items():
            # counts는 {'TP': TP, 'TN': TN, 'FN': FN, 'FP': FP} 형태입니다.
            TP = counts['TP']
            TN = counts['TN']
            FP = counts['FP']
            FN = counts['FN']

            # 성능 메트릭 계산
            Recall = round(TP / (TP + FN) if (TP + FN) != 0 else 0 , 4)
            Precision = round(TP / (TP + FP) if (TP + FP) != 0 else 0 , 4)
            F1_SCORE = round(2 * (Precision * Recall) / (Precision + Recall) if (Precision + Recall) != 0 else 0 , 4)
            Accuracy = round((TP + TN) / (TP + TN + FP + FN) if (TP + TN + FP + FN) != 0 else 0 , 4)

            # 추가적인 계산
            TP_FN = TP + FN
            FP_TN = FP + TN
            Total = TP + FP + TN + FN

            # 결과를 딕셔너리로 저장
            result = {
                'Alpha': alpha,
                'TP': TP,
                'FP': FP,
                'FN': FN,
                'TN': TN,
                'TP + FN': TP_FN,
                'FP + TN': FP_TN,
                'Total': Total,
                'Precision': Precision,
                'Recall': Recall,
                'Accuracy': Accuracy,
                'F1_Score': F1_SCORE
            }

            # 결과를 리스트에 추가
            data_list.append(result)

            # 각 인원 수별로 별도의 DataFrame 생성 및 저장
            self.performance_dict[num_people] = pd.DataFrame([result])

            # 결과를 self.Q에 저장
            self.Q[num_people] = result

        # 전체 결과를 DataFrame으로 생성
        self.performance_df = pd.DataFrame(data_list, columns=[
            'Alpha', 'TP', 'FP', 'FN', 'TN',
            'TP + FN', 'FP + TN', 'Total',
            'Precision', 'Recall', 'Accuracy', 'F1_Score'
        ])


    
    # 결과를 출력하는 메서드 (선택 사항)
    def print_performance(self):
        # 전체 결과 출력
        print("전체 결과:")
        print(self.performance_df)

        # 각 인원 수별 결과 출력
        for num_people, df in self.performance_dict.items():
            print(f"\n인원 수 {num_people}명에 대한 성능 메트릭:")
            print(df)
if __name__ == "__main__":

    from glob import glob
    from natsort import  natsorted

    name = 'cv2_diff_test/raw_seat/11'

    number = name.split('/')[-1]
    distort_images = natsorted(glob(os.path.join(f'{name}','*.jpg')))

    distort_images = natsorted(glob(os.path.join(f"cv2_diff_test/front",'**',"*.jpg"),recursive=True))
    if not distort_images:
        raise FileExistsError
    
    import natsort
    import glob
    #distort_images = natsorted(glob.glob(os.path.join("cv2_diff_test/problem" , "*.jpg")))
    #distort_images = natsort.natsorted(glob.glob(os.path.join('cv2_diff_test/front/4.1' ,'**','*.jpg'), recursive=True))
    #distort_images = natsort.natsorted(glob.glob(os.path.join('cv2_diff_test/front/krri1' ,'**','*.jpg'), recursive=True))
    #distort_images = distort_images[355 :]  # 혼잡한 이미지
    #distort_images = natsort.natsorted(glob.glob(os.path.join('cv2_diff_test/front/krri1' ,'**','*.jpg'), recursive=True))
    #distort_images = distort_images[270 : ]
    errors = []
    import gc

    range_ = np.linspace(0, 1 ,11)
    #range_ = [0.8]
    box_p  = [7,8,10,11,12,13] # 6명
    #box_p = [7,8,11,12] # 4명
    #new_people = # [9,5,6] # 9명 추가 
    #box_p.extend(new_people)
    #box_p = [1,2,3,4]
    #box_p.sort()
    #range_ = list(reversed(range_))
    # range_ = list(map (lambda x: round(x,2) , range_))

    #range_ = [0.5 , 0.6 , 0.7  , 0.8  , 0.9 , 1]
    #range_ = [0 , 0.1]
    print(range_)
    print(box_p)
    print(f"{len(box_p)} 명 수행..")
    input("============ continue press key ! ==================")

    c = YOLODetector(distort_images, number, alpha=range_ , box_p= box_p)
    c.run_all

    # import matplotlib.pyplot as plt
    # plt.figure(figsize=(12, 12))
    # plt.plot(range_ , errors, label='Error over alpha' , color = 'red')
    # plt.xlabel('Alpha')
    # plt.ylabel('Error')
    # plt.title('Error vs Alpha')
    # plt.legend()
    # plt.show()
