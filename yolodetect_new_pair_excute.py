import os
import cv2
import numpy as np
from ultralytics import YOLO
from tqdm.auto import  tqdm
import os
import fish_map
import box_module
import logging
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%H:%M:%S'
)

logging.info("Yolo Detect run ...")

import time
import functools
def timing_decorator(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        current_alpha = args[1]
        print("Model : ",args[0].model.model_name)
        print("ALPHA: ", current_alpha)
        print(f"Function '{func.__name__}' started at {time.strftime('%H:%M:%S', time.localtime(start_time))}")
        
        result = func(*args, **kwargs)
        
        end_time = time.time()
        print(f"Function '{func.__name__}' ended at {time.strftime('%H:%M:%S', time.localtime(end_time))}")
        print(f"Total execution time: {end_time - start_time:.4f} seconds")

        return result
    return wrapper


class YOLODetector:
    def __init__(self, f1 , number , alpha , box_p ,model_path='yolo8x.pt'):
        self.number = number
        self.distortion_list = f1
        self.model = YOLO(model_path).to("cuda")
        self.empty_count = 0
        self.false_alarm = 0
        self.not_same_source = 0

        self.alpha = alpha

        self.box_people = box_p

        self.image_error_list = [
            'cv2_diff_test/front/4.1/image_0229.jpg',
            'cv2_diff_test/front/4.1/image_0237.jpg' , 
        ]

        self.total_people = 0
        self.TP = 0 
        self.TN = 0
        self.FP = 0
        self.FN = 0

        self.TP_2 = 0 
        self.TN_2 = 0
        self.FP_2 = 0
        self.FN_2 = 0

    @property
    def run_all(self):
        for alpha_value in self.alpha:
            self.run(alpha_value)
        
    def append_data_to_excel(self, dataFrame, alpha_value ,sheet_name = f"sheet_1"):
            import openpyxl
            from openpyxl import load_workbook, Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            import pandas as pd
            
            file_path = 'resulted.xlsx'
            try:
                excel_file = openpyxl.load_workbook(file_path)
                if sheet_name in excel_file.sheetnames:
                    excel_ws = excel_file[sheet_name]
                    if alpha_value == self.alpha[0]:
                        excel_ws.append(dataFrame.columns.tolist())

                else:
                    excel_ws = excel_file.create_sheet(sheet_name)
                    excel_ws.append(dataFrame.columns.tolist())

            except:
                wb = Workbook()
                excel_ws = wb.active
                excel_ws.title = sheet_name

                excel_ws.append(dataFrame.columns.tolist())

                wb.save(file_path)

                excel_file = openpyxl.load_workbook(file_path)
                excel_ws = excel_file[sheet_name]

            for r in dataFrame.itertuples(index= False , name=None):
                print(r)
                excel_ws.append(r)

            excel_file.save(file_path)
    
    def nmx_box_to_cv2_loc(self , boxes):
        x1 , y1 , w, h = boxes
        x2 = x1 + w
        y2 = y1 + h

        return [x1 , y1 , x2 , y2]

    def apply_nms(self, boxes, scores, iou_threshold=0.6):
        indices = cv2.dnn.NMSBoxes(boxes, scores, 0.5, iou_threshold)
        if isinstance(indices, list) and len(indices) > 0:
            return [boxes[i[0]] for i in indices]
        elif isinstance(indices, np.ndarray) and indices.size > 0:
            return [boxes[i] for i in indices.flatten()]
        else:
            return []
        
    @timing_decorator
    def run(self , alpha_value):
        for index , i in tqdm(enumerate(self.distortion_list),total=len(self.distortion_list)):
            image_name = i
            distort_image   = cv2.imread(i , cv2.IMREAD_COLOR)

            if distort_image is not None:
                h, w, _ = list(map(int, distort_image.shape))

                # Add black padding around the distorted image
                black = np.zeros(((int(w - h) // 2), w, 3), dtype=np.uint8)
                frame_new = cv2.vconcat([black, distort_image])
                frame_new = cv2.vconcat([frame_new, black])

                # Recalculate the height and width after padding
                h, w, _ = list(map(int, frame_new.shape))

                # Apply fisheye_to_plane_info mapping on the newly padded image
                undisort_image = np.array(fish_map.fisheye_to_plane_info(frame_new, h, w, 180, 90, 640, 0, 0))
                undisort_image = cv2.resize(undisort_image, (640, 640), interpolation=cv2.INTER_LANCZOS4)
            

            result = self.model(undisort_image , verbose = False , classes = 0)

            boxes = []
            scores = []

            for res in result:
                for box in res.boxes:
                    if int(box.cls) == 0:
                        x1 , y1 , w , h = box.xyxy[0].tolist()
                        score = box.conf
                        boxes.append([int(x1), int(y1), int(w - x1), int(h - y1)])
                        scores.append(float(score.detach().cpu().numpy()))

            nmx_boxes = self.apply_nms(boxes , scores)
            nmx_boxes = list(map(self.nmx_box_to_cv2_loc , nmx_boxes))
            if not nmx_boxes:
                self.empty_count +=1
                continue


            answer , answer_location , ans_img =  box_module.second(nmx_boxes , undisort_image.copy() , Answer= True , ALPHA = 1 )
            compare_match , compare_location  , compare_img = box_module.second(nmx_boxes , undisort_image.copy() , Answer= False, ALPHA = alpha_value)
           # compare_match_2 , compare_location_2  , compare_img_2 = box_module.second(nmx_boxes , undisort_image.copy() , Answer= False, ALPHA = self.alpha[1])


            self.Metrix(answer_location , compare_location , box_p , flag = True )

            self.total_people += len(compare_location.keys())     # 전체명수    
            answer_location = dict(sorted(answer_location.items()))
            compare_location = dict(sorted(compare_location.items()))

            if answer != list(answer_location.keys()) or compare_match != list(compare_location.keys()):
                print(answer)
                print(answer_location)
                print(compare_match)
                print(compare_location)
                raise ValueError
            


            stop_point = False
            if image_name in self.image_error_list:
                if image_name == 'cv2_diff_test/front/4.1/image_0229.jpg':
                    answer = [5 , 7 , 9 , 12 ]
                else:
                    answer = [5 , 7 , 9 , 12 ,13 ]
                if sorted(answer) != sorted(compare_location.keys()):
                    self.false_alarm += 1
                    stop_point = True
            else:
                if len(answer_location) == len(compare_location):  # 길이가 같은 경우
                    for answer_key in answer_location:
                        if answer_key not in compare_location:  # 키가 다르면 오경보 처리
                            self.false_alarm += 1
                            stop_point = True
                            break
                        else:  # 키가 같다면 값 비교
                            if answer_location[answer_key] != compare_location[answer_key]:
                                self.not_same_source += 1
                                stop_point = True
                                break
                else:  # 길이가 다르면 오경보 처리
                    self.false_alarm += 1
                    stop_point = True

            if stop_point:
                print('===================================================')
                print(f"{image_name}")
                print(f"{answer_location}")
                print(f"{compare_location}")
                print(f"오경보 : {self.false_alarm}")
                print(f"출처다름 : {self.not_same_source}")
            

            # cv2.namedWindow("Answer" , cv2.WINDOW_NORMAL)
            # cv2.namedWindow("compare", cv2.WINDOW_NORMAL)
            # cv2.imshow("Answer" , ans_img)
            # cv2.imshow("compare",compare_img)
            # cv2.waitKey(0)
            # cv2.destroyAllWindows()

            
        print("전체 사람 수 : ",self.total_people)
        print("TP : ", self.TP)
        print("FN : " , self.FN)
        print("FP : ",self.FP)
        print("TN : ", self.TN)

        
        
        Criterion , Confusion = self.Performeance_Metrix()
        Confusion.insert(0,round(alpha_value,1))
        Criterion = list(map (lambda x : round(x,4), Criterion))
        PR , RE , F1, ACC = Criterion
        Confusion.extend(Criterion)
        dataframe = pd.DataFrame([Confusion] , columns=['Alpha','TP','FP','FN','TN','TP + FN','FP + TN','TP+FP+TN+FN','Precison','Recall','Accuracy','F1'])
        self.append_data_to_excel(dataframe , alpha_value)
        print("Alpha  : ",self.alpha)
        print("Recall : {} Precision : {} F1 : {} ACC : {}".format(RE , PR , F1 , ACC))

    def Metrix(self, answer_location, compare_location , box_p , flag ):

        mapper = [None for i in range(13)] 
        ans_key = answer_location.keys() # 1 ~ 13 { 9 , 10 ,13}
        compare_key = compare_location.keys()  # 1 ~ 13
        ans_key = list(map(lambda x : x- 1 , ans_key))
        compare_key = list(map(lambda x : x- 1 , compare_key))
        for i in range(13):
            if i in ans_key and i in compare_key:
                mapper[i] = 'TP'
            elif i not in ans_key and i not in compare_key:
                mapper[i] = 'TN'
            elif i in ans_key and i not in compare_key:
                mapper[i] = 'FN'
            elif i not in ans_key and i in compare_key:  
                mapper[i] = 'FP'


        for compare_key, value in compare_location.items():
            adjust_key = compare_key - 1
            try:
                answer_value = answer_location[compare_key]
                if value != answer_value:
                    mapper[adjust_key] = 'FP' 

            except KeyError:
                if adjust_key not in ans_key: 
                    mapper[adjust_key] = 'FP'

        if box_p:  # empty full 상태
            user = box_p # 정상 좌석 번호 입력
            mapping_ = list(map(lambda x : x-1 ,user )) # [6,7,9,10,11,12]
            for idx , i in enumerate(mapper):
                if idx not in mapping_:
                        mapper[idx] = None
        
        # 결과 계산
        TP = FP = FN = TN = 0

        if flag:
            TP_attr, TN_attr, FN_attr, FP_attr = 'TP', 'TN', 'FN', 'FP'
        else:
            TP_attr, TN_attr, FN_attr, FP_attr = 'TP_2', 'TN_2', 'FN_2', 'FP_2'

        for i in mapper:
            if i == None:
                continue
            if i == 'TP':
                TP += 1
                setattr(self, TP_attr, getattr(self, TP_attr) + 1)
            elif i == 'TN':
                TN += 1
                setattr(self, TN_attr, getattr(self, TN_attr) + 1)
            elif i == 'FN':
                FN += 1
                setattr(self, FN_attr, getattr(self, FN_attr) + 1)
            elif i == 'FP':
                FP += 1
                setattr(self, FP_attr, getattr(self, FP_attr) + 1)

        if (TP + FP + FN + TN) != len(user):
            print(f"TP: {TP}, FP: {FP}, FN: {FN}, TN: {TN}")
            raise ValueError(f"Total does not equal {len(user)}.")
           
    def Performeance_Metrix(self):

        TP = self.TP
        TN = self.TN
        FP = self.FP
        FN = self.FN
        self.TP = self.TN = self.FP = self.FN = 0
        TN = TN - 556
        FN = FN + 556

        Recall = TP / (TP + FN) if (TP + FN) != 0 else 0
        Precision = TP / (TP + FP) if (TP + FP) != 0 else 0
        F1_SCORE = 2 * (Precision * Recall) / (Precision + Recall) if (Precision + Recall) != 0 else 0
        Accuracy = (TP + TN) / (TP + TN + FP + FN) if (TP + TN + FP + FN) != 0 else 0



        ASSESSMENT_CRITERIA = [Precision, Recall, Accuracy, F1_SCORE]
        CONFUSION_METRIX = [TP, FP, FN, TN , TP + FN , FP + TN , TP + FP + TN + FN]
        return ASSESSMENT_CRITERIA , CONFUSION_METRIX
        
if __name__ == "__main__":

    from glob import glob
    from natsort import  natsorted

    name = 'cv2_diff_test/raw_seat/11'

    number = name.split('/')[-1]
    distort_images = natsorted(glob(os.path.join(f'{name}','*.jpg')))

    distort_images = natsorted(glob(os.path.join(f"cv2_diff_test/front",'**',"*.jpg"),recursive=True))
    if not distort_images:
        raise FileExistsError
    
    import natsort
    import glob
    #distort_images = natsorted(glob(os.path.join("cv2_diff_test/problem" , "*.jpg")))
    #distort_images = natsort.natsorted(glob.glob(os.path.join('cv2_diff_test/front/1.1' ,'**','*.jpg'), recursive=True))
    #distort_images = natsort.natsorted(glob.glob(os.path.join('cv2_diff_test/front/krri1' ,'**','*.jpg'), recursive=True))
    #distort_images = natsort.natsorted(glob.glob(os.path.join('cv2_diff_test/front/krri1' ,'**','*.jpg'), recursive=True))
    #distort_images = distort_images[270 : ]
    errors = []
    import gc

    range_ = np.linspace(0, 1 ,11)
    box_p  = [7,8,10,11,12,13]
    new_people = [9,5,6]
    box_p.extend(new_people)
    box_p.sort()
    # range_ = list(reversed(range_))
    # range_ = list(map (lambda x: round(x,2) , range_))

    #range_ = [0.5 , 0.6 , 0.7  , 0.8  , 0.9 , 1]
    #range_ = [0 , 0.1]
    print(range_)
    print(box_p)
    input("============ continue press key ! ==================")

    c = YOLODetector(distort_images, number, alpha=range_ , box_p= box_p)
    c.run_all

    # import matplotlib.pyplot as plt
    # plt.figure(figsize=(12, 12))
    # plt.plot(range_ , errors, label='Error over alpha' , color = 'red')
    # plt.xlabel('Alpha')
    # plt.ylabel('Error')
    # plt.title('Error vs Alpha')
    # plt.legend()
    # plt.show()
