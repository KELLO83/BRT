import os
import cv2
import numpy as np
from ultralytics import YOLO
from tqdm.auto import  tqdm
from natsort import  natsort
import matplotlib.pyplot as plt
import matplotlib
import os
import fish_map
import pandas as pd
import openpyxl
import matplotlib
import re
from super_gradients.training import models
class YOLODetector:
    def __init__(self, f1 , model_path='yolov8x.pt'):
        self.distortion_list = f1
        self.model = YOLO(model_path).to("cuda")
        self.plot_save_dir = 'plot_front'
        
        self.mask_image = cv2.imread('cv2_diff_test/IMAGE_3RGB_MASK_USE/24.jpg' , cv2.IMREAD_GRAYSCALE)
        self.mask_cordinate = np.where(self.mask_image >=1)
        self.mask_cordinate = list(zip(self.mask_cordinate[0] , self.mask_cordinate[1]))

        self.dataFrame = pd.DataFrame(columns=['distort','undistort'])


    def append_data_to_excel(self, sheet_name = f"sheet_Rm"):
        import openpyxl
        from openpyxl import load_workbook, Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        colums_avg = self.dataFrame.mean()
        colums_sum = self.dataFrame.sum()

        file_path = 'Yolo_compare.xlsx'
        try:
            excel_file = openpyxl.load_workbook(file_path)
            if sheet_name in excel_file.sheetnames:
                excel_ws = excel_file(sheet_name)
            else:
                excel_ws = excel_file.create_sheet(sheet_name)
                excel_ws.append(self.dataFrame.columns.tolist())

        except:
            wb = Workbook()
            excel_ws = wb.active
            excel_ws.title = sheet_name

            excel_ws.append(self.dataFrame.columns.tolist())

            wb.save(file_path)

            excel_file = openpyxl.load_workbook(file_path)
            excel_ws = excel_file[sheet_name]

        for r in self.dataFrame.itertuples(index= False):
            excel_ws.append(r)


        # excel_file.save(file_path)
        print('=======================================')
        print("colums_sum : \n",colums_sum)
        print("colums_avg : \n", colums_avg)
        print('=======================================')

            
    def apply_nms(self, boxes, scores, iou_threshold=0.4):
        indices = cv2.dnn.NMSBoxes(boxes, scores, 0.5, iou_threshold)
        if isinstance(indices, list) and len(indices) > 0:
            return [boxes[i[0]] for i in indices]
        elif isinstance(indices, np.ndarray) and indices.size > 0:
            return [boxes[i] for i in indices.flatten()]
        else:
            return []

    def run(self):
        total_d = len(self.distortion_list)

        std_dis = 0
        std_und = 0

        noc = []
        unknown_detction = 0
        for index , i in tqdm(enumerate(self.distortion_list),total=total_d):
            distort_image   = cv2.imread(i , cv2.IMREAD_COLOR)
            # cv2.namedWindow("t",cv2.WINDOW_NORMAL)
            # cv2.imshow("t",distort_image)
            # cv2.waitKey(0)

            # print(i) # cv2_diff_test/front2/1.1_image_0041.jpg
            # T = i.split('/')[-1]
            # SCEN_INDEX = T.find('_')
            # SCEN = T[ : SCEN_INDEX]
            # match = re.search(r'\d{4}', T)
            # Number = match.group()
            # Number = int(Number)

            # people_exist = 0
        
            # if SCEN == '1.1':
            #     people_exist = 7

            # elif SCEN == '4.1':
            #     people_exist = 7

            # elif SCEN == 'krri1':
            #     if 4 <= Number <= 125:
            #         people_exist = 2
            #     elif 265 <= Number <= 351:
            #         people_exist = 5
            #     elif 352 <= Number <= 400:
            #         people_exist = 5
            
            # else:
            #     noc.append(os.path.basename(i))
            #     continue



            if distort_image is not None:
                h, w, _ = list(map(int, distort_image.shape))

                # Add black padding around the distorted image
                black = np.zeros(((int(w - h) // 2), w, 3), dtype=np.uint8)
                frame_new = cv2.vconcat([black, distort_image])
                frame_new = cv2.vconcat([frame_new, black])

                # Recalculate the height and width after padding
                h, w, _ = list(map(int, frame_new.shape))

                # Apply fisheye_to_plane_info mapping on the newly padded image
                undisort_image = np.array(fish_map.fisheye_to_plane_info(frame_new, h, w, 180, 90, 600, 0, 0))


            #distort_image = cv2.resize(distort_image , dsize=(600,600))
            #undisort_image = cv2.resize(undisort_image , dsize=(600,600))

            distort_result = self.model(distort_image , verbose=False , classes = 0)
            undisort_result = self.model(undisort_image , verbose = False , classes = 0)

            distort_box_image , dis_count , is_seat_dis  = self.generate_box(distort_result , distort_image , color = 'red'  , sig =True)
            undistort_box_image , und_count  , is_seat_und  = self.generate_box(undisort_result , undisort_image , color = 'blue' , sig = False)

            if is_seat_dis:
                std_dis +=1
            
            if is_seat_und:
                std_und +=1

            new_row = {'distort' : [dis_count] , 'undistort' : [und_count]}
            new_pd = pd.DataFrame(new_row)
            self.dataFrame =  pd.concat([self.dataFrame , new_pd] , ignore_index= True)


            cv2.namedWindow("d",cv2.WINDOW_NORMAL)
            cv2.namedWindow("und" , cv2.WINDOW_NORMAL)
            cv2.imshow('d' , distort_box_image)
            cv2.imshow('und', undistort_box_image)
            cv2.waitKey(0)

            SAVE_DIR = 'TEST'
            os.makedirs(SAVE_DIR , exist_ok= True)
            name = os.path.join(SAVE_DIR , os.path.basename(i))
            #cv2.imwrite(name , undistort_box_image)
 
            
                
            distort_box_image = cv2.cvtColor(distort_box_image , cv2.COLOR_BGR2RGB)
            undistort_box_image = cv2.cvtColor(undistort_box_image  , cv2.COLOR_BGR2RGB)
            distort_box_image = cv2.resize(distort_box_image , dsize = (600,600))
            undistort_box_image = cv2.resize(undistort_box_image , dsize=(600,600))
            # plt.figure(figsize=(8, 8))
            # plt.subplot(1, 2, 1)  
            # plt.title(f"{dis_count}")
            # plt.imshow(distort_box_image)


            # plt.subplot(1, 2, 2)  
            # plt.title(f"{und_count}")
            # plt.imshow(undistort_box_image)


            # # Display both images
            # save_name = os.path.join(self.plot_save_dir , f"{index}.jpg")
            # os.makedirs(self.plot_save_dir , exist_ok=True)
            # plt.savefig(save_name)
            # plt.close()


        print("Noc : \n",noc)
        print("Unknown Detection : " , unknown_detction)
        print("왜곡이미지 착석판별 :" , std_dis)
        print("펴진이미지 착석판별 :" , std_und)


        self.append_data_to_excel()

    def generate_box(self , results , img , color , sig ):
        boxes = []
        scores = []
        colors = {
            'red' : (0,0,255),
            'blue' : (255,0,0),
        }

        # if sig : # 11번좌석
        #     focus_area = [681,255,789,390]
        
        # else:
        #     focus_area = [161,138 , 222 , 222]

        if sig : # 5번좌석
            focus_area = [1000,388,1154 , 674]
        
        else:
            focus_area = [318 ,227 , 399 , 377]
        
        x1 , y1 , x2 , y2 = focus_area
        
        # y_range = abs(y2 - y1)
        # y_pad = y_range // 4

        # middle_y = (y2 + y1) // 2
        
        # y1 = middle_y - y_pad
        # y2 = middle_y + y_pad

        # focus_area = x1 , y1 , x2 , y2
        
        # if sig :  # 4번좌석
        #     focus_area = [1188 , 488 , 1339 , 729]
        # else:
        #     focus_area = [420 , 271 , 515, 410]

        # if sig: # 24번좌석 
        #     foucs_area = [385 , 122 , 437 , 196]
        
        # else:
        #     foucs_area = [373 , 154 , 419 , 223]


        # if sig : # 23번좌석
        #     foucs_area = [314 , 120 , 393 , 205]
        # else:
        #     foucs_area = [312 , 156 , 381, 227]


        # PAD = 60
        # if sig: # 17번좌석
        #     foucs_area = [419 , 245 , 515, 496]
        #     x1 , y1 ,x2 , y2 = foucs_area
        #     middle_x = (x1 + x2) // 2
        #     middle_y = (y1 + y2) // 2
        #     y1 = middle_y - PAD
        #     y2 = middle_y + PAD

        #     foucs_area[1] = y1
        #     foucs_area[3] = y2

        # else:
        #     foucs_area = [402 , 255, 503, 465]

        #     x1 , y1 ,x2 ,y2 = foucs_area
        #     middle_x = (x1 + x2) // 2
        #     middle_y = (y1 + y2) // 2
        #     y1 = middle_y - PAD
        #     y2 = middle_y + PAD
        #     foucs_area[1] = y1
        #     foucs_area[3] = y2
        

        for result in results:
            for box in result.boxes:
                if int(box.cls) == 0:
                    x1 , y1 ,x2 , y2 = box.xyxy[0].tolist()
                    score = box.conf
                    boxes.append([int(x1), int(y1), int(x2 - x1), int(y2 - y1)])
                    scores.append(float(score))

        nmx_boxes = self.apply_nms(boxes , scores)

        box_count = 0
        is_exist = False
        for box in nmx_boxes:
            x1, y1, w, h = box
            center_x, center_y = x1 + w // 2, y1 + h // 2
            
            cv2.rectangle(img , (focus_area[0],focus_area[1]), (focus_area[2],focus_area[3]) , (255,102,255) , 1)
            cv2.rectangle(img , (x1 , y1) ,(x1+w , y1+h) , (0,255,0) , 1)
            cv2.circle(img, (center_x, center_y), 5, colors[color], -1)
            # 좌표를 이미지에 작게 표시
            cv2.putText(img, f'({center_x}, {center_y})', 
                        (center_x, center_y), 
                        cv2.FONT_HERSHEY_SIMPLEX, 
                        0.3, 
                        (255, 255, 255), 
                        1)

            E = self.is_area(center_x , center_y ,focus_area)

            if E :
                is_exist = True

            box_count += 1            

        return img , box_count , is_exist
    
    def is_area(self , center_x , center_y , f):
        x1 , y1 , x2 , y2 = f

        if x1 <= center_x <= x2 and y1 <= center_y <= y2:
            return True
        else :
            return False
        


if __name__ == "__main__":
    from glob import glob
    from natsort import  natsorted
    distort_images = natsorted(glob(os.path.join('cv2_diff_test/raw_seat/5','*.jpg')))
    #distort_images = natsorted(glob(os.path.join('cv2_diff_test/front' , '**','*.jpg'),recursive=True))
    #distort_images = natsorted(glob(os.path.join('cv2_diff_test/front2' , '*.jpg')))
    undisort_images = natsorted(glob(os.path.join('Undisort_krri/_camera8_image_raw', "*.jpg")))

    c = YOLODetector( distort_images )
    c.run()