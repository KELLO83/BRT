import os
import cv2
import numpy as np
from ultralytics import YOLO
from tqdm.auto import  tqdm
from natsort import  natsort
import matplotlib.pyplot as plt
import matplotlib
import os
import fish_map
import pandas as pd
import openpyxl
import matplotlib
import re
from super_gradients.training import models
from PIL import Image
import generat_box_module
import copy
class YOLODetector:
    def __init__(self, f1 , number ,model_path='yolov8x-seg.pt'):
        self.number = number
        self.distortion_list = f1
        self.model = YOLO(model_path).to("cuda")
        self.plot_save_dir = 'plot_front'
        
        self.mask_image = cv2.imread('cv2_diff_test/IMAGE_3RGB_MASK_USE/24.jpg' , cv2.IMREAD_GRAYSCALE)
        self.mask_cordinate = np.where(self.mask_image >=1)
        self.mask_cordinate = list(zip(self.mask_cordinate[0] , self.mask_cordinate[1]))

        self.dataFrame = pd.DataFrame(columns=['distort','undistort'])


    def append_data_to_excel(self, sheet_name = f"sheet_Rm"):
        import openpyxl
        from openpyxl import load_workbook, Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        colums_avg = self.dataFrame.mean()
        colums_sum = self.dataFrame.sum()

        file_path = 'Yolo_compare.xlsx'
        try:
            excel_file = openpyxl.load_workbook(file_path)
            if sheet_name in excel_file.sheetnames:
                excel_ws = excel_file(sheet_name)
            else:
                excel_ws = excel_file.create_sheet(sheet_name)
                excel_ws.append(self.dataFrame.columns.tolist())

        except:
            wb = Workbook()
            excel_ws = wb.active
            excel_ws.title = sheet_name

            excel_ws.append(self.dataFrame.columns.tolist())

            wb.save(file_path)

            excel_file = openpyxl.load_workbook(file_path)
            excel_ws = excel_file[sheet_name]

        for r in self.dataFrame.itertuples(index= False):
            excel_ws.append(r)


        # excel_file.save(file_path)
        print('=======================================')
        print("colums_sum : \n",colums_sum)
        print("colums_avg : \n", colums_avg)
        print('=======================================')

            
    def apply_nms(self, boxes, scores, iou_threshold=0.6):
        indices = cv2.dnn.NMSBoxes(boxes, scores, 0.5, iou_threshold)
        if isinstance(indices, list) and len(indices) > 0:
            return [boxes[i[0]] for i in indices]
        elif isinstance(indices, np.ndarray) and indices.size > 0:
            return [boxes[i] for i in indices.flatten()]
        else:
            return []

    def run(self):
        total_d = len(self.distortion_list)

        std_dis = 0
        std_und = 0

        unknown_detction = 0
        False_ALRAM = 0
        for index , i in tqdm(enumerate(self.distortion_list),total=total_d):
            distort_image   = cv2.imread(i , cv2.IMREAD_COLOR)


            if distort_image is not None:
                h, w, _ = list(map(int, distort_image.shape))

                # Add black padding around the distorted image
                black = np.zeros(((int(w - h) // 2), w, 3), dtype=np.uint8)
                frame_new = cv2.vconcat([black, distort_image])
                frame_new = cv2.vconcat([frame_new, black])

                # Recalculate the height and width after padding
                h, w, _ = list(map(int, frame_new.shape))

                # Apply fisheye_to_plane_info mapping on the newly padded image
                undisort_image = np.array(fish_map.fisheye_to_plane_info(frame_new, h, w, 180, 90, 600, 0, 0))

    

            cv2.namedWindow("d",cv2.WINDOW_NORMAL)
            cv2.namedWindow("und" , cv2.WINDOW_NORMAL)

            distort_result = self.model(distort_image , verbose=False , classes = 0)
            undisort_result = self.model(undisort_image , verbose = False , classes = 0)
            undisort_result_copy = copy.deepcopy(undisort_result)
            #distort_box_image , dis_count , is_seat_dis  = self.generate_box(distort_result , distort_image , number= self.number , sig =True , crop=True)
            #undistort_box_image , und_count  , is_seat_und  = self.generate_box(undisort_result , undisort_image , number= self.number , sig = False)
            distort_box_image , dis_count , is_seat_dis , b1 = generat_box_module.call_generate_box_final(undisort_result , undisort_image , number=self.number)
            cv2.imshow('d' , distort_box_image)
            undistort_box_image , und_count , is_seat_und , b2 = generat_box_module.call_generate_box_only(undisort_result_copy , undisort_image , number=self.number)

            # if not is_seat_dis:
            #     cv2.namedWindow("d",cv2.WINDOW_NORMAL)
            #     cv2.namedWindow("und" , cv2.WINDOW_NORMAL)
            #     cv2.imshow('d' , distort_box_image)
            #     cv2.imshow('und', undistort_box_image)
            #     cv2.waitKey(0)

            if is_seat_dis and not is_seat_und:
                unknown_detction +=1 

            elif is_seat_dis and is_seat_und:
                flag = generat_box_module.compare(b1 , b2)

                if not flag:
                    False_ALRAM += 1

            elif not is_seat_dis and is_seat_und:
                False_ALRAM += 1
            
            if is_seat_dis:
                std_dis +=1
            if is_seat_und:
                std_und +=1

            new_row = {'distort' : [dis_count] , 'undistort' : [und_count]}
            new_pd = pd.DataFrame(new_row)
            self.dataFrame =  pd.concat([self.dataFrame , new_pd] , ignore_index= True)

      


            cv2.imshow('und', undistort_box_image)
            cv2.waitKey(0)

            #distort_box_image , undistort_box_image = list(map(lambda x : cv2.cvtColor(x , cv2.COLOR_BGR2RGB) , [distort_box_image , undistort_box_image]))

            # SAVE_DIR = 'TEST'
            # os.makedirs(SAVE_DIR , exist_ok= True)
            # name = os.path.join(SAVE_DIR , os.path.basename(i))
            # cv2.imwrite(name , undistort_box_image)
            
                
            distort_box_image = cv2.cvtColor(distort_box_image , cv2.COLOR_BGR2RGB)
            undistort_box_image = cv2.cvtColor(undistort_box_image  , cv2.COLOR_BGR2RGB)
            distort_box_image = cv2.resize(distort_box_image , dsize = (600,600))
            undistort_box_image = cv2.resize(undistort_box_image , dsize=(600,600))
            # plt.figure(figsize=(8, 8))
            # plt.subplot(1, 2, 1)  
            # plt.title(f"{dis_count}")
            # plt.imshow(distort_box_image)


            # plt.subplot(1, 2, 2)  
            # plt.title(f"{und_count}")
            # plt.imshow(undistort_box_image)


            # # Display both images
            # save_name = os.path.join(self.plot_save_dir , f"{index}.jpg")
            # os.makedirs(self.plot_save_dir , exist_ok=True)
            # plt.savefig(save_name)
            # plt.close()

        print('FALSE ALRAM : ',False_ALRAM)
        print("Unknown Detection : " , unknown_detction)
        print("왜곡이미지 착석판별 {} {:.2f}%" .format(std_dis , round(std_dis / total_d , 2) * 100))
        print("펴진이미지 착석판별 {:.2f} , {}%" .format(std_und , round(std_und / total_d , 2) * 100))

        self.append_data_to_excel()

    
if __name__ == "__main__":
    from glob import glob
    from natsort import  natsorted

    name = 'cv2_diff_test/raw_seat/7'

    number = name.split('/')[-1]
    distort_images = natsorted(glob(os.path.join(f'{name}','*.jpg')))
    if not distort_images:
        raise FileExistsError
    

    #distort_images = natsorted(glob(os.path.join('cv2_diff_test/raw_seat/4','*.jpg')))
    #distort_images = natsorted(glob(os.path.join('cv2_diff_test/front' , '**','*.jpg'),recursive=True))
    #distort_images = natsorted(glob(os.path.join('cv2_diff_test/front2' , '*.jpg')))
    #undisort_images = natsorted(glob(os.path.join('Undisort_krri/_camera8_image_raw', "*.jpg")))

    c = YOLODetector( distort_images , number)
    c.run()